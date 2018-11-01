import datetime
import random
from subprocess import Popen

import openpyxl
import peewee
import xlrd
from telebot import TeleBot
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, InlineKeyboardMarkup, \
    InlineKeyboardButton

from classes import Users, Reminder, Events, Bot_settings
from telegramcalendar import create_calendar, number_keyboard

Popen("send_messages.py", shell=True)
bot = TeleBot("727398167:AAFa6E7ZjjieCbpqpJhe9CDu_OCazY3vnKs")
database = peewee.SqliteDatabase("database.db")
telebot = Bot_settings()


def get_id():
    """

    Функция определяет свободный id при создании
    нового мероприятия.

    """
    try:
        user_id = 0
        for i in Events.select():
            if i.id > user_id:
                user_id = i.id
        user_id += 1
    except Events.DoesNotExist:
        user_id = 0
    return user_id


def event_invite(msg):
    """

    Функция осуществляет рассылку всем пользователям,
    у которых выбранна данная категория мероприятия.

    """
    chosen_event = Events.select().where((Events.creator == msg.chat.id) & (Events.status == 4)).get()
    for i in Users.select():
        if msg.chat.id != i.id and i.fun.find(chosen_event.fun) + 1:
            telebot.keyboard = InlineKeyboardMarkup()
            telebot.keyboard.add(
                InlineKeyboardButton(text='Хочу пойти', callback_data='ev_invite' + str(chosen_event.id)))
            bot.send_message(i.id, text='✉\nНовое мероприятие! ' + '\n' +
                                        '⌚Время: ' + str(chosen_event.time) + '\n' +
                                        '📅Дата: ' + str(chosen_event.date) + '\n' +
                                        '📄Описание:' + chosen_event.text,
                             reply_markup=telebot.keyboard)
    chosen_event.status = 0
    chosen_event.save()


def event_create_step1(msg):
    """

    Функция записывает категорию, выбранную пользователем и создаёт
    новое поле в таблице Events для дальнейшего заполнения.

    """
    get_calendar(msg)
    bot.send_message(msg.chat.id,
                     text='Ваша категория добавлена, теперь напишите краткое описание вашего мероприятия',
                     reply_markup=telebot.keyboard)
    chosen_event = Events.create(
        id=get_id(),
        creator=msg.chat.id,
        date=datetime.date(1, 1, 1),
        time=datetime.time(0, 0, 0),
        text='NULL',
        count=-1,
        fun=msg.text,
        address='NULL',
        members='',
        status=1
    )
    chosen_event.save()


def event_create_step2(msg):
    """

    Функция определяет по статусу, на какой стадии
    находится создание мероприятия.
    Статус "1" - заполненяется ячейка категории мероприятия и описания
    Статус "2" - заполненяется ячейка времени и даты мероприятия
    Статус "3" - заполняется геолокация мероприятия
    Статус "4" - заполняется ячейка количества человек, производится рассылка
    приглашений, статус мероприятия становится "Успешно создано"


    """
    try:
        chosen_event = Events.select().where((Events.count == -1) & (Events.creator == msg.chat.id)).get()
        if chosen_event.status == 1:
            if telebot.date == datetime.date(1, 1, 1):
                bot.send_message(msg.chat.id, text='Ты забыл ввести дату!')
                return
            chosen_event.date = telebot.date
            chosen_event.text = msg.text
            bot.send_message(msg.chat.id, text='Укажите время мероприятия (формат HH:MM)')
            chosen_event.status = 2
            chosen_event.save()
        elif chosen_event.status == 2:
            try:
                chosen_event.time = datetime.time(int(msg.text[0:2]), int(msg.text[3:5]))
                bot.send_message(msg.chat.id, text='Отправьте геолокацию с адресом мероприятия')
                chosen_event.status = 3
                chosen_event.save()
            except ValueError:
                bot.send_message(msg.chat.id, text='Проверь правильность ввода времени')
    except Events.DoesNotExist:
        bot.send_message(msg.chat.id, text='Произошла ошибка при создании мероприятия')


def event_create(msg):
    """

    Функция определяет, на какой стадии создания мероприятия
    находится пользователь

    """
    if msg.text + '\n' in telebot.lines:
        event_create_step1(msg)
    else:
        event_create_step2(msg)


def event_list(msg):
    """

    Функция извлекает из базы данных все мероприятия, в которых
    пользователь является администратором или участником.

    """
    events = Events.select().where((Events.creator == msg.chat.id) & (Events.status == 0))
    if len(events) > 0:
        bot.send_message(msg.chat.id, text='Ваши мероприятия:', reply_markup=telebot.keyboard)
    count_admin = 1
    for i in events:
        event_information(msg, i, count_admin, 1)
        count_admin += 1
    count_member = 1
    for i in Events.select():
        if i.members.find(str(msg.chat.id)) != -1:
            if count_member == 1:
                bot.send_message(msg.chat.id, text='Мероприятия, на которые вы идёте:', reply_markup=telebot.keyboard)
            event_information(msg, i, count_member, 0)
            count_member += 1
    if count_admin == 1 and count_member == 1:
        bot.send_message(msg.chat.id, text='У вас нет активных мероприятий', reply_markup=telebot.keyboard)


def event_information(msg, chosen_event, number, is_creator):
    """

    Функция извлекает из базы данных переданное мероприятие,
    добавляет к нему функции "Покинуть, показать подробнее",
    преобразует в читабельный вид.

    """
    telebot.keyboard = InlineKeyboardMarkup()
    text = str(number) + ') ' + chosen_event.text + '\n' + 'Дата: ' + str(chosen_event.date) + '\n'
    telebot.keyboard.add(InlineKeyboardButton(text='Подробнее...', callback_data='info_' + str(chosen_event.id)))
    if is_creator:
        telebot.keyboard.add(InlineKeyboardButton(text='Удалить', callback_data='del_' + str(chosen_event.id)))
    else:
        telebot.keyboard.add(InlineKeyboardButton(text='Покинуть', callback_data='leave_' + str(chosen_event.id)))
    bot.send_message(msg.chat.id, text=text, reply_markup=telebot.keyboard)


def event(msg):
    """

    Функция перенаправляет ответ пользователя на разные ветки
    создания или просмотра мероприятий.

    """
    if telebot.action[msg.chat.id] == 'event':
        if msg.text == 'Создать мероприятие':
            telebot.keyboard = ReplyKeyboardMarkup()
            for i in telebot.lines:
                telebot.keyboard.add(i)
            bot.send_message(msg.chat.id, text='Выберите категорию вашего мероприятия:', reply_markup=telebot.keyboard)
            telebot.keyboard = ReplyKeyboardRemove()
            telebot.action[msg.chat.id] = 'event_create'
        elif msg.text == 'Посмотреть список моих мероприятий':
            telebot.keyboard = ReplyKeyboardRemove()
            event_list(msg)
            telebot.action[msg.chat.id] = 'answer'
    elif telebot.action[msg.chat.id] == 'event_create':
        event_create(msg)


def fun_adding(msg):
    """

    Функция извлекает доступные категории мероприятий из
    файла и предоставляет пользователю выбор.

    """
    telebot.keyboard = InlineKeyboardMarkup()
    k = 0
    for i in telebot.lines:
        telebot.keyboard.add(InlineKeyboardButton(text=i, callback_data='fun_' + str(k)))
        k += 1
    telebot.keyboard.add(InlineKeyboardButton(text="Завершить", callback_data='fun_end'))
    bot.send_message(msg.chat.id, text='Выбери категории развлечений, по которым ты хотел бы получать приглашения',
                     reply_markup=telebot.keyboard)
    telebot.action[msg.chat.id] = 'fun_add'


def fun_removing(msg):
    """

    Функция извлекает доступные категории мероприятий
    из файла и предоставляет пользователю возможность
    отписаться от ненужных категорий.

    """
    telebot.keyboard = InlineKeyboardMarkup()
    k = 0
    for i in telebot.lines:
        telebot.keyboard.add(InlineKeyboardButton(text=i, callback_data='fun_' + str(k)))
        k += 1
    telebot.keyboard.add(InlineKeyboardButton(text="Завершить", callback_data='fun_end'))
    bot.send_message(msg.chat.id, text='Удали развлечения из этого списка', reply_markup=telebot.keyboard)
    telebot.action[msg.chat.id] = 'fun_remove'


def fun(msg):
    """

    Функция перенаправляет пользователя на ветки
    выбора и удаления категорий.

    """
    if telebot.action[msg.chat.id] == 'fun':
        if msg.text == 'Добавить категорию':
            fun_adding(msg)
        elif msg.text == 'Удалить категорию':
            fun_removing(msg)


def review(msg):
    """

    Функция октрывает файл и записывает в него отзыв
    пользователя.

    """
    if telebot.action[msg.chat.id] != "review":
        bot.send_message(msg.chat.id, text="Напиши мне отзыв в следующем сообщении и я его передам разработчику!",
                         reply_markup=telebot.keyboard)
        telebot.action[msg.chat.id] = "review"
    else:
        bot.send_message(msg.chat.id, text="Записал твой отзыв, спасибо!", reply_markup=telebot.keyboard)
        rb = xlrd.open_workbook('review.xlsx')
        sheet = rb.sheet_by_index(0)
        wb = openpyxl.load_workbook('review.xlsx')
        sheet1 = wb['Лист1']
        sheet1.cell(row=sheet.nrows + 1, column=1).value = msg.chat.id
        sheet1.cell(row=sheet.nrows + 1, column=2).value = msg.from_user.first_name
        sheet1.cell(row=sheet.nrows + 1, column=3).value = msg.from_user.last_name
        sheet1.cell(row=sheet.nrows + 1, column=4).value = msg.text
        wb.save('review.xlsx')
        telebot.action[msg.chat.id] = 'answer'


def find_friend(msg):
    if not access(msg):
        return
    """

    Функция извлекает из базы данных ключевые слова
    пользователя и сверяет их с ключевыми словами
    других пользователей. При совпадении происходит
    рассылка и обмен контактами.

    """
    user = Users.get(Users.id == msg.chat.id)
    hobbies = list(user.hobbies.split())
    bot.send_message(msg.chat.id, text='Выполняю поиск...')
    for i in hobbies:
        for j in Users.select():
            hobbies_friend = list(j.hobbies.split())
            if i in hobbies_friend and j.id != user.id:
                bot.send_message(j.id,
                                 text='Я нашёл тебе друга!\n🙂 {} {}\n'
                                      '📊 Репутация: {}\n'
                                      '📱 Телефон: {}'.format(
                                                              user.first_name,
                                                              user.last_name,
                                                              str(user.reputation),
                                                              user.telephone))
                bot.send_message(msg.chat.id,
                                 text='Я нашёл тебе друга!\n🙂 {} {}\n'
                                      '📊Репутация: {}\n'
                                      '📱Телефон: {}'.format(
                                                             j.first_name,
                                                             j.last_name,
                                                             str(j.reputation),
                                                             j.telephone))
                return
    bot.send_message(msg.chat.id, text='Друг не найден(')


def memory(msg):
    """

    Функция извлекает время, дату и описание из сообщения
    и создаёт новую ячейку в базе данных. Если время или
    дата введены не верно - функция попросит пользователя
    повторить попытку.

    """
    try:
        time = datetime.time(int(msg.text[0:2]), int(msg.text[3:5]))
        text = msg.text[6:]
        reminder = Reminder.create(
            id=msg.chat.id,
            text=text,
            date=telebot.date,
            time=time
        )
        reminder.save()
        bot.send_message(msg.chat.id, text="Хорошо! Обязательно тебе напомню!")
        telebot.action[msg.chat.id] = 'answer'
        telebot.date = ''
    except ValueError:
        bot.send_message(msg.chat.id, text="Проверь правильность ввода данных! Не забудь указать дату!")
        return


def value_reg(msg):
    """

    Функция первоначальной регистрации пользователя, в базе данных
    остаётся номер телефона пользователя, имя, фамилия, увлечения.

    """
    user = Users.get(Users.id == msg.chat.id)
    telebot.keyboard = ReplyKeyboardRemove()
    if 'Я прочитал и ознакомился с правилами' == msg.text or telebot.action[msg.chat.id] == 'reg_hobbies' or \
            telebot.action[msg.chat.id] == 'reg_end':
        if telebot.action[msg.chat.id] == 'reg_telephone':
            telebot.keyboard = ReplyKeyboardMarkup()
            bot.send_message(msg.chat.id, text="Отлично! Теперь тебе нужно внести данные, это не займёт много времени!",
                             reply_markup=telebot.keyboard)
            telebot.keyboard.add(
                KeyboardButton("Отправить номер телефона", request_contact=True)
            )
            bot.send_message(msg.chat.id, text='Отправь мне свой номер телефона', reply_markup=telebot.keyboard)
            telebot.action[msg.chat.id] = 'reg_hobbies'
        elif telebot.action[msg.chat.id] == 'reg_hobbies':
            bot.send_message(msg.chat.id,
                             text='Записал твой номер. Теперь отметь хэштэги по своим интересам, чтобы другим людям.'
                                  'было проще найти тебя.', reply_markup=telebot.keyboard)
            user.country = msg.text
            user.save()
            telebot.action[msg.chat.id] = 'reg_end'
        elif telebot.action[msg.chat.id] == 'reg_end':
            bot.send_message(msg.chat.id, text='Записал твои хобби, спасибо за регистрацию!')
            user.hobbies += msg.text
            user.save()
            fun_adding(msg)
    elif msg.text == 'Я отказываюсь предоставлять доступ к моим данным':
        telebot.keyboard = ReplyKeyboardRemove()
        bot.send_message(msg.chat.id, text='В таком случае вы не будете иметь доступ к коммандам\n/events, '
                                           '/find_friend, /fun',
                         reply_markup=telebot.keyboard)


def actions(msg):
    """

    Функция определяет текущее действия пользователя
    и перенаправляет на нужную ветку.

    """
    if 'fun' in telebot.action[msg.chat.id]:
        fun(msg)
    if 'weather_reg' in telebot.action[msg.chat.id]:
        weather_reg(msg)
    elif 'reg' in telebot.action[msg.chat.id]:
        value_reg(msg)
    elif telebot.action[msg.chat.id] == 'answer':
        bot.send_message(msg.chat.id, text=answer(msg), reply_markup=telebot.keyboard)
    elif telebot.action[msg.chat.id] == 'review':
        review(msg)
    elif telebot.action[msg.chat.id] == 'memory':
        memory(msg)
    elif 'event' in telebot.action[msg.chat.id]:
        event(msg)


def weather_reg(msg):
    """

    Функция извлекает данные о времени и записывает в базу данных

    """
    user = Users.get(Users.id == msg.chat.id)
    if telebot.action[msg.chat.id] == 'weather_reg':
        if msg.text == 'Да' or msg.text == 'да':
            user.weather = 1
            user.save()
            telebot.keyboard = ReplyKeyboardRemove()
            bot.send_message(msg.chat.id, text='В какое время ты бы хотел получать уведомления?',
                             reply_markup=telebot.keyboard)
            telebot.action[msg.chat.id] = 'weather_reg1'
        if msg.text == 'Нет' or msg.text == 'нет':
            user.weather = -1
            user.save()
            telebot.keyboard = ReplyKeyboardRemove()
            bot.send_message(msg.chat.id, text="Хорошо, как скажешь)", reply_markup=telebot.keyboard)
    else:
        try:
            user.weather_time = datetime.time(int(msg.text[0:2]), int(msg.text[3:5]))
            bot.send_message(msg.chat.id, text='Хорошо! Буду тебя предупреждать каждый день в ' + msg.text,
                             reply_markup=telebot.keyboard)
            telebot.action[msg.chat.id] = 'answer'
            user.save()
        except Users.DoesNotExist or TypeError:
            bot.send_message(msg.chat.id, text='Не правильный ввод, повтори ещё раз!', reply_markup=telebot.keyboard)


def weather(msg, latitude, longitude):
    """

    Функция отправляет пользователю актуальную погоду и
    предлагает включить уведомление на каждый день.

    """
    bot.send_message(msg.chat.id, text=telebot.weather_text(latitude, longitude))
    user = Users.get(Users.id == msg.chat.id)
    if user.weather == 0:
        telebot.keyboard = ReplyKeyboardMarkup()
        telebot.keyboard.add(
            KeyboardButton('Да'),
            KeyboardButton('Нет'))
        bot.send_message(msg.chat.id, text="Хочешь, чтобы я сообщал тебе погоду каждый день?",
                         reply_markup=telebot.keyboard)
        telebot.action[msg.chat.id] = 'weather_reg'
        return


def hello(msg):
    """

    Функция проверяет наличие пользователя в базе данных.
    Если его нет - создаётся новая ячейка, в которую записывается
    id чата пользователя.
    Если он есть - бот отправляет приветствие.

    """
    try:
        user = Users.get(Users.id == msg.chat.id)
        bot.send_message(msg.chat.id,
                         text='Приветствую вас, {}!\nЕсли вам нужна помощь, используйте команду /help'.format(
                             user.first_name))
    except Users.DoesNotExist:
        bot.send_message(msg.chat.id, text="Привет! Рад с тобой познакомиться!")
        first_name = msg.from_user.first_name
        last_name = msg.from_user.last_name
        if msg.from_user.first_name is None:
            first_name = 'Unnamed'
        if msg.from_user.last_name is None:
            last_name = ' '
        chosen_user = Users.create(id=msg.chat.id,
                                   telephone='NULL',
                                   hobbies='',
                                   first_name=first_name,
                                   last_name=last_name,
                                   reputation=0,
                                   latitude=0.0,
                                   longitude=0.0,
                                   weather=0,
                                   weather_time=datetime.time(0, 0, 0),
                                   fun=''
                                   )
        chosen_user.save()
        registration(msg)


@bot.message_handler(commands=['reg'])
def registration(msg):
    telebot.keyboard = ReplyKeyboardMarkup()
    telebot.keyboard.add(
        KeyboardButton("Я прочитал и ознакомился с правилами"),
        KeyboardButton("Я отказываюсь предоставлять доступ к моим данным"))
    bot.send_message(msg.chat.id,
                     text='📃Для доступа ко всем функциям бота '
                          'тебе нужно ознакомиться с условиями конфиденциальности:\n'
                          '🔸Бот будет использовать ваш номер телефона, никнейм, имя и фамилию.\n',
                     reply_markup=telebot.keyboard)
    telebot.action[msg.chat.id] = 'reg_telephone'


def answer(msg):
    """

    Функция работает с текстом, присланным пользователем,
    распознаёт слова и перенаправляет на ветки для
    последующих команд.

    """
    text = msg.text.lower()
    for i in telebot.words.welcome:
        i = i.lower()
        if i.find(text) != -1:
            return telebot.words.welcome[random.randint(0, len(telebot.words.welcome) - 1)]
    for i in telebot.words.leave:
        i = i.lower()
        if i.find(text) != -1:
            return telebot.words.leave[random.randint(0, len(telebot.words.leave) - 1)]
    if text.find("как") + 1 and text.find("дела") + 1:
        telebot.keyboard = ReplyKeyboardMarkup()
        telebot.keyboard.add(
            KeyboardButton("Плохо" + telebot.emoji.pictures['грусть']),
            KeyboardButton("Хорошо" + telebot.emoji.pictures['улыбка']),
            KeyboardButton("Отлично" + telebot.emoji.pictures['улыбка1']))
        return "У меня всё хорошо, а у вас?"
    elif text.find('!#!') + 1:  # функция используется только администратором
        for i in Users.select():
            if i.id != msg.chat.id:
                bot.send_message(i.id, text='БОГ: '+msg.text[4:])
        bot.send_message(msg.chat.id, text='Ваше сообщение отправлено всем пользователям!')
    elif text.find("плохо") + 1:
        telebot.keyboard = ReplyKeyboardRemove()
        return "Надеюсь, что в скором времени будет хорошо:)" + telebot.emoji.pictures['подмигивание']
    elif text.find("хорошо") + 1:
        telebot.keyboard = ReplyKeyboardRemove()
        return "Рад за вас!"
    elif text.find("отлично") + 1:
        telebot.keyboard = ReplyKeyboardRemove()
        return "Это просто прекрасно!"
    elif text.find("погода") + 1 or text.find("погоду") + 1 or text.find("погоде") + 1 or text.find("погодой") + 1:
        receive_weather(msg)
    elif text.find("отзыв") + 1:
        review(msg)
    elif text.find('репутация') + 1:
        receive_reputation(msg)
    elif text.find('поменя') + 1 and text.find('время') + 1 and text.find('уведомлени') + 1:
        receive_change_weather(msg)
    elif text.find('найди') + 1 and text.find('друга'):
        find_friend(msg)
    elif text.find('напомин') + 1 or text.find('напомни') + 1:
        receive_memory(msg)
    elif text.find('развлечен') + 1:
        receive_fun(msg)
    elif text.find('мероприяти') + 1:
        receive_event(msg)
    else:
        bot.send_message(msg.chat.id, text='Я тебя не понимаю')


def access(msg):
    chosen_user = Users.get(Users.id == msg.chat.id)
    if chosen_user.telephone == 'NULL':
        bot.send_message(msg.chat.id, text='У вас нет доступа к этой комманде', reply_markup=telebot.keyboard)
        return False
    else:
        return True


def get_calendar(msg):
    now = datetime.datetime.now()  # Current date
    chat_id = msg.chat.id
    telebot.date = (now.year, now.month)
    telebot.current_shown_dates[chat_id] = telebot.date  # Saving the current date in a dict
    markup = create_calendar(now.year, now.month)
    bot.send_message(msg.chat.id, "Пожалуйста, выберете дату", reply_markup=markup)


"""

Данный блок кода принимает команды пользователя

"""


@bot.message_handler(commands=['number'])
def send_keyboard(msg):
    markup = number_keyboard()
    bot.send_message(msg.chat.id, text="Укажите время: ", reply_markup=markup)


@bot.message_handler(commands=['delete'])
def delete(msg):
    """

    Функция используется только администратором

    """
    chosen_user = Users.get(Users.id == msg.chat.id)
    chosen_user.delete_instance()
    bot.send_message(msg.chat.id, text='Вы были успешно удалены из базы данных')


@bot.message_handler(commands=['weather'])
def receive_weather(msg):
    user = Users.get(Users.id == msg.chat.id)
    if user.latitude == 0 or user.longitude == 0:
        telebot.keyboard = ReplyKeyboardMarkup()
        telebot.keyboard.add(
            KeyboardButton("Отправить геолокацию", request_location=True)
        )
        bot.send_message(msg.chat.id, text='Прости, но я не знаю твоей геолокации:(', reply_markup=telebot.keyboard)
        return
    else:
        weather(msg, user.latitude, user.longitude)


@bot.message_handler(commands=['events'])
def receive_event(msg):
    if not access(msg):
        return
    telebot.keyboard = ReplyKeyboardMarkup()
    telebot.keyboard.add(
        KeyboardButton('Создать мероприятие'),
        KeyboardButton('Посмотреть список моих мероприятий')
    )
    bot.send_message(msg.chat.id, text='Что вы хотите сделать?', reply_markup=telebot.keyboard)
    telebot.action[msg.chat.id] = 'event'
    telebot.keyboard = ReplyKeyboardRemove()


@bot.message_handler(commands=['find_friend'])
def receive_friend(msg):
    find_friend(msg)


@bot.message_handler(commands=['fun'])
def receive_fun(msg):
    if not access(msg):
        return
    telebot.keyboard = ReplyKeyboardMarkup()
    telebot.keyboard.add(
        KeyboardButton('Добавить категорию'),
        KeyboardButton('Удалить категорию')
    )
    bot.send_message(msg.chat.id, text='Что вы хотите сделать?', reply_markup=telebot.keyboard)
    telebot.action[msg.chat.id] = 'fun'
    actions(msg)


@bot.message_handler(commands=['change_weather'])
def receive_change_weather(msg):
    telebot.keyboard = ReplyKeyboardMarkup()
    telebot.keyboard.add(
        KeyboardButton('Да'),
        KeyboardButton('Нет'))
    bot.send_message(msg.chat.id, text="Хочешь, чтобы я сообщел тебе погоду каждый день?",
                     reply_markup=telebot.keyboard)
    telebot.action[msg.chat.id] = 'weather_reg'


@bot.message_handler(commands=['memory'])
def receive_memory(msg):
    get_calendar(msg)
    telebot.action[msg.chat.id] = 'memory'


@bot.message_handler(commands=['review'])
def receive_review(msg):
    review(msg)


@bot.message_handler(commands=['reputation'])
def receive_reputation(msg):
    user = Users.get(Users.id == msg.chat.id)
    if 2 > user.reputation > -2:
        bot.send_message(msg.chat.id, text='Твоя репутация: ' + str(user.reputation) + ' (нейтральная)')
    elif int(user.reputation) > 2:
        bot.send_message(msg.chat.id, text='Твоя репутация: ' + str(
            user.reputation) + ' (к тебе хорошо относятся другие пользователи)')
    else:
        bot.send_message(msg.chat.id, text='Твоя репутация: ' + str(
            user.reputation) + ' (к тебе плохо относятся другие пользователи)')


@bot.message_handler(commands=['start'])
def start(msg):
    hello(msg)


@bot.message_handler(commands=['help'])
def information(msg):
    bot.send_message(msg.chat.id, text='/weather - Узнать погоду по вашему местоположению\n' +
                                       '/events - Создать/Удалить/Узнать ваши мероприятия\n' +
                                       '/find_friend - Найти друга со схожими интересами\n' +
                                       '/reg - Пройти регистрацию\n' +
                                       '/fun - Редактировать категории мероприятий\n' +
                                       '/change_weather - Изменить время отправки погоды\n' +
                                       '/reputation - Посмотреть свою репутацию\n' +
                                       '/help - Посмотреть список комманд\n' +
                                       '/start - Начать работу\n' +
                                       '/cancel - Отменить последнее действие\n' +
                                       '/review - Оставить отзыв\n' +
                                       '/memory - Добавить напоминание\n')


@bot.message_handler(commands=['cancel'])
def cancel(msg):
    telebot.keyboard = ReplyKeyboardRemove()
    bot.send_message(msg.chat.id, text='Действие отменено', reply_markup=telebot.keyboard)
    telebot.action[msg.chat.id] = 'answer'
    try:
        chosen_event = Events.select().where(Events.count == -1).get()
        chosen_event.delete_instance()
        chosen_event.save()
    except Events.DoesNotExist:
        return


@bot.message_handler(content_types=["text"])
def receive_message(msg):
    if msg.chat.id not in telebot.action.keys():
        telebot.action[msg.chat.id] = 'answer'
    try:
        chosen_user = Users.get(Users.id == msg.chat.id)
        chosen_user.save()
    except Users.DoesNotExist:
        hello(msg)

    actions(msg)


@bot.message_handler(content_types='contact')
def phone(msg):
    chosen_user = Users.get(Users.id == msg.chat.id)
    chosen_user.telephone = msg.contact.phone_number
    chosen_user.save()
    if telebot.action[msg.chat.id] == 'reg_hobbies':
        telebot.keyboard = ReplyKeyboardRemove()
        value_reg(msg)


@bot.message_handler(content_types='location')
def location(msg):
    if telebot.action[msg.chat.id] != 'event_create':
        telebot.keyboard = ReplyKeyboardRemove()
        user = Users.get(Users.id == msg.chat.id)
        user.latitude = msg.location.latitude
        user.longitude = msg.location.longitude
        user.save()
        bot.send_message(msg.chat.id, text='Записал твою геолокацию, спасибо!', reply_markup=telebot.keyboard)
        weather(msg, user.latitude, user.longitude)
    else:
        chosen_event = Events.select().where((Events.count == -1) & (Events.creator == msg.chat.id)).get()
        chosen_event.address = str(msg.location.latitude) + ',' + str(msg.location.longitude)
        chosen_event.status = 4
        chosen_event.count = 0
        chosen_event.save()
        telebot.action[msg.chat.id] = 'answer'
        telebot.date = datetime.datetime(1, 1, 1)
        bot.send_message(msg.chat.id, text='Мероприятие успешно создано!')
        event_invite(msg)


@bot.callback_query_handler(func=lambda call: call.data[0:13] == 'calendar-day-')
def get_day(call):
    chat_id = call.message.chat.id
    saved_date = telebot.current_shown_dates.get(chat_id)
    if saved_date is not None:
        day = call.data[13:]
        telebot.date = datetime.date(int(saved_date[0]), int(saved_date[1]), int(day))
        bot.answer_callback_query(call.id, text="Дата выбрана")
        if telebot.action[call.message.chat.id] == 'memory':
            bot.send_message(call.message.chat.id,
                             text='Напиши время и само напоминание в формате HH:MM ')

    else:
        bot.answer_callback_query(call.id, text="Ошибка ввода даты")


@bot.callback_query_handler(func=lambda call: call.data == 'next-month')
def next_month(call):
    chat_id = call.message.chat.id
    saved_date = telebot.current_shown_dates.get(chat_id)
    if saved_date is not None:
        year, month = saved_date
        month += 1
        if month > 12:
            month = 1
            year += 1
        date = (year, month)
        telebot.current_shown_dates[chat_id] = date
        markup = create_calendar(year, month)
        bot.edit_message_text("Please, choose a date", call.from_user.id, call.message.message_id, reply_markup=markup)
        bot.answer_callback_query(call.id, text="")
    else:
        pass


@bot.callback_query_handler(func=lambda call: call.data == 'previous-month')
def previous_month(call):
    chat_id = call.message.chat.id
    saved_date = telebot.current_shown_dates.get(chat_id)
    if saved_date is not None:
        year, month = saved_date
        month -= 1
        if month < 1:
            month = 12
            year -= 1
        date = (year, month)
        telebot.current_shown_dates[chat_id] = date
        markup = create_calendar(year, month)
        bot.edit_message_text("Please, choose a date", call.from_user.id, call.message.message_id, reply_markup=markup)
        bot.answer_callback_query(call.id, text="")
    else:
        # Do something to inform of the error
        pass


@bot.callback_query_handler(func=lambda call: call.data == 'ignore')
def ignore(call):
    bot.answer_callback_query(call.id, text="")


"""

Данный блок принимает информацию о нажатых callback кнопках 

"""


@bot.callback_query_handler(func=lambda call: 'fun' in call.data)
def fun_call(call):
    if call.data == 'fun_end':
        fun_call_end(call)
        return
    chosen_fun = telebot.lines[int(call.data[4:])]
    user = Users.get(Users.id == call.message.chat.id)
    if telebot.action[call.message.chat.id] == 'fun_add':
        fun_call_add(call, chosen_fun, user)
    if telebot.action[call.message.chat.id] == 'fun_remove':
        fun_call_remove(call, chosen_fun, user)


def fun_call_end(call):
    telebot.keyboard = ReplyKeyboardRemove()
    bot.send_message(call.message.chat.id, text='Изменения внесены ✅', reply_markup=telebot.keyboard)
    telebot.action[call.message.chat.id] = 'answer'


def fun_call_add(call, chosen_fun, user):
    if not chosen_fun[:len(chosen_fun) - 1] in user.fun:
        user.fun += ' ' + chosen_fun[:len(chosen_fun) - 1]
        user.save()
        bot.answer_callback_query(call.id, text="Развлечение добавлено")
    else:
        bot.answer_callback_query(call.id, text="Развлечение уже добавлено")


def fun_call_remove(call, chosen_fun, user):
    if chosen_fun[:len(chosen_fun) - 1] in user.fun:
        user.fun = user.fun.replace(chosen_fun[:len(chosen_fun) - 1], '')
        user.fun = user.fun.replace('  ', ' ')
        user.save()
        bot.answer_callback_query(call.id, text="Развлечение удалено")
    else:
        bot.answer_callback_query(call.id, text="Такого развлечения нет")


@bot.callback_query_handler(func=lambda call: 'ev' in call.data)
def event_call(call):
    if call.data[0:9] == 'ev_invite':
        try:
            chosen_user = Users.get(Users.id == call.message.chat.id)
            chosen_event = Events.get(Events.id == int(call.data[9:]))
            if chosen_event.members.find(str(call.message.chat.id)) == -1:
                telebot.keyboard = InlineKeyboardMarkup()
                telebot.keyboard.add(
                    InlineKeyboardButton(text='Принять', callback_data='ev_accept' + str(chosen_event.id) + ':' + str(
                        call.message.chat.id)))
                telebot.keyboard.add(
                    InlineKeyboardButton(text='Отклонить', callback_data='ev_reject' + str(chosen_event.id) + ':' + str(
                        call.message.chat.id)))
                bot.send_message(chosen_event.creator,
                                 text='✉\nНа ваше мероприятие записался человек!\n🙂 {} {}\n'
                                      '📊 Репутация: {}\n📱 Телефон: {}'.format(
                                                                                chosen_user.first_name,
                                                                                chosen_user.last_name,
                                                                                str(chosen_user.reputation),
                                                                                chosen_user.telephone),
                                 reply_markup=telebot.keyboard)
                bot.edit_message_text("Ваша заявка отправлена", call.from_user.id, call.message.message_id)
        except Users.DoesNotExist or Events.DoesNotExist:
            bot.send_message(call.message.chat.id, text="К сожалению, мероприятия больше не существует")
    elif call.data[0:9] == 'ev_accept':
        event_id = int(call.data[9:call.data.find(':')])
        user_id = int(call.data[call.data.find(':') + 1:])
        chosen_event = Events.get(Events.id == event_id)
        if chosen_event.members.find(str(user_id)) == -1:
            telebot.keyboard = InlineKeyboardMarkup()
            url = InlineKeyboardButton(text="Адрес", url="https://www.google.ru/maps/place/" + chosen_event.address)
            telebot.keyboard.add(url)
            bot.send_message(user_id,
                             text='✉\nВаша заявка одобрена!\n⌚ Время: {}\n📅 Дата: {}\n📄 Описание: {}'.format(
                                 str(chosen_event.time), str(chosen_event.date), chosen_event.text),
                             reply_markup=telebot.keyboard)
            chosen_event.count += 1
            chosen_event.members += str(user_id) + ' '
            chosen_event.save()
            bot.send_message(chosen_event.creator, text='Пользователю отправлена полная информация о мероприятии')
        else:
            bot.send_message(chosen_event.creator, text='Этот пользователь уже приглашён на мероприятие')

    elif call.data[0:9] == 'ev_reject':
        event_id = int(call.data[9:call.data.find(':')])
        creator = int(call.data[call.data.find(':') + 1:])
        chosen_event = Events.get(Events.id == event_id)
        bot.send_message(creator, text='✉\nВаша заявка на мероприятие: "' + chosen_event.text + '" отклонена!')


@bot.callback_query_handler(func=lambda call: 'info_' in call.data)
def event_info(call):
    try:
        event_id = call.data[5:]
        chosen_event = Events.get(Events.id == event_id)
        admin = Users.get(Users.id == chosen_event.creator)
        telebot.keyboard = InlineKeyboardMarkup()
        url = InlineKeyboardButton(text="Адрес", url="https://www.google.ru/maps/place/" + chosen_event.address)
        telebot.keyboard.add(url)
        text = '📄 Описание: {}\n⌚ Время: {}\n📅 Дата: {}'.format(
            chosen_event.text,
            str(chosen_event.time),
            str(chosen_event.date))
        text1 = '🙂 Создатель: {}\n{}\n📱 Телефон: {}\n📊 Репутация: {}'.format(
            admin.first_name,
            admin.last_name,
            admin.telephone,
            str(admin.reputation))
        bot.send_message(call.message.chat.id, text=text, reply_markup=telebot.keyboard)
        bot.send_message(call.message.chat.id, text=text1)
        text2 = 'Участники:' + '\n'
        for members in chosen_event.members.split():
            chosen_user = Users.get(Users.id == int(members))
            text2 += '🙂 {} {}\n📱 Телефон: {}'.format(
                chosen_user.first_name,
                chosen_user.last_name,
                chosen_user.telephone)
        if len(text2) > 11:
            bot.send_message(call.message.chat.id, text=text2)
    except Events.DoesNotExist or Users.DoesNotExist:
        pass


@bot.callback_query_handler(func=lambda call: 'del_' in call.data)
def event_delete(call):
    try:
        event_id = call.data[4:]
        chosen_event = Events.get(Events.id == event_id)
        for i in list(chosen_event.members.split()):
            bot.send_message(int(i), text='✉К сожалению создатель удалил мероприятие "' + chosen_event.text + '"')
        bot.send_message(chosen_event.creator, text='✉Мероприятие успешно удалено')
        chosen_event.delete_instance()
        chosen_event.save()
    except Events.DoesNotExist:
        pass


@bot.callback_query_handler(func=lambda call: 'leave_' in call.data)
def event_delete(call):
    telebot.keyboard = InlineKeyboardMarkup()
    event_id = call.data[6:]
    user_id = str(call.message.chat.id)
    try:
        chosen_event = Events.get(Events.id == event_id)
        if chosen_event.members.find(str(user_id)) != -1:
            chosen_event.members = chosen_event.members.replace(str(user_id), ' ')
            chosen_event.members = chosen_event.members.replace('  ', ' ')
            chosen_event.count -= 1
            telebot.keyboard.add(InlineKeyboardButton(text='Подробнее...', callback_data='info_' +
                                                                                         str(chosen_event.id)))
            bot.send_message(int(chosen_event.creator), text='✉К сожалению, ваше мероприятие покинул человек!',
                             reply_markup=telebot.keyboard)
            chosen_event.save()
        else:
            bot.answer_callback_query(call.id, text='Вы уже покинули это мероприятие')
    except Events.DoesNotExist:
        bot.send_message(call.message.chat.id, text='Мероприятия не существует')


@bot.callback_query_handler(func=lambda call: 'rep+' in call.data)
def rep_positive(call):
    user = Users.get(Users.id == int(call.data[5:]))
    user.reputation += 1
    bot.edit_message_text("*пользователь оценён*", call.from_user.id, call.message.message_id)
    user.save()


@bot.callback_query_handler(func=lambda call: 'rep-' in call.data)
def rep_positive(call):
    user = Users.get(Users.id == int(call.data[5:]))
    user.reputation -= 1
    bot.edit_message_text("*пользователь оценён*", call.from_user.id, call.message.message_id)
    user.save()


@bot.callback_query_handler(func=lambda call: 'number_' in call.data)
def send_keyboard(call):
    markup = number_keyboard()
    if len(call.data) == 8:
        if len(telebot.time) == 2:
            telebot.time += ':' + call.data[7:]
        else:
            telebot.time += call.data[7:]
        bot.edit_message_text("Укажите время: " + telebot.time, call.from_user.id, call.message.message_id,
                              reply_markup=markup)
    elif call.data == "number_back":
        if len(telebot.time) == 4:
            telebot.time = telebot.time[:3]
        else:
            telebot.time = telebot.time[:len(telebot.time) - 1]
        bot.edit_message_text("Укажите время: " + telebot.time, call.from_user.id, call.message.message_id,
                              reply_markup=markup)
    elif call.data == "number_done":
        bot.edit_message_text("Вы выбрали время: " + telebot.time, call.from_user.id, call.message.message_id,
                              reply_markup=markup)
    elif call.data == "number_clear":
        telebot.time = ''
        bot.edit_message_text("Укажите время: " + telebot.time, call.from_user.id, call.message.message_id,
                              reply_markup=markup)


bot.polling(none_stop=True)
